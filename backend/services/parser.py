"""
Парсер листа ОТЧЕТ_МБ. ПРОТЕСТИРОВАН — НЕ ПЕРЕПИСЫВАТЬ.

Выход parse_report():
{
    "filename": str,
    "dates": [date, ...],
    "units": {
        "Название установки": {
            "inputs":  DataFrame (product|plan_month|plan_day|YYYY-MM-DD_meas|YYYY-MM-DD_recon|...),
            "outputs": DataFrame,
            "summary": {
                "consumed":     {"measured": [float...], "reconciled": [float...]},
                "produced":     {"measured": [float...], "reconciled": [float...]},
                "imbalance":    {"measured": [float...], "reconciled": [float...]},
                "imbalance_rel":{"measured": [float...], "reconciled": [float...]},
            }
        }
    }
}
"""
import openpyxl
import pandas as pd
from datetime import datetime, date
from typing import Dict, List
import os, sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from config import (
    SHEET_NAME, HEADER_DATE_ROW, DATA_START_COL,
    COL_UNIT, COL_FLOW_DIR, COL_PRODUCT, COL_PLAN_MONTH, COL_PLAN_DAY
)


def parse_report(filepath: str) -> Dict:
    wb = openpyxl.load_workbook(filepath, data_only=True, keep_vba=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Лист '{SHEET_NAME}' не найден. Есть: {wb.sheetnames}")
    ws = wb[SHEET_NAME]
    dates = _parse_dates(ws)
    units = _parse_unit_blocks(ws, dates)
    wb.close()
    return {"filename": os.path.basename(filepath), "dates": dates, "units": units}


def _parse_dates(ws) -> List[date]:
    dates = []
    col = DATA_START_COL
    while col <= ws.max_column:
        val = ws.cell(row=HEADER_DATE_ROW, column=col).value
        if val is None:
            break
        if isinstance(val, datetime):
            dates.append(val.date())
        elif isinstance(val, date):
            dates.append(val)
        elif isinstance(val, str):
            for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
                try:
                    dates.append(datetime.strptime(val, fmt).date())
                    break
                except ValueError:
                    continue
        col += 2
    return dates


def _parse_unit_blocks(ws, dates: List[date]) -> Dict:
    units = {}
    num_days = len(dates)
    cur_unit = None
    cur_dir = None
    cur_prods = []

    for row_idx in range(1, ws.max_row + 1):
        f = ws.cell(row=row_idx, column=COL_UNIT).value
        g = ws.cell(row=row_idx, column=COL_FLOW_DIR).value
        h = ws.cell(row=row_idx, column=COL_PRODUCT).value

        if f and f != 'Потоки':
            if cur_unit and cur_dir and cur_prods:
                _save(units, cur_unit, cur_dir, cur_prods, dates)
                cur_prods = []
            cur_unit = str(f).strip()
            if cur_unit not in units:
                units[cur_unit] = {
                    "inputs": None, "outputs": None,
                    "summary": {k: {"measured": [], "reconciled": []}
                                for k in ("consumed", "produced", "imbalance", "imbalance_rel")}
                }

        if g in ('Входящие', 'Исходящие'):
            if cur_unit and cur_dir and cur_prods:
                _save(units, cur_unit, cur_dir, cur_prods, dates)
                cur_prods = []
            cur_dir = g

        if g in ('Потреблено', 'Вырабатано', 'Дебаланс', 'Дебаланс отн.'):
            if cur_unit and cur_dir and cur_prods:
                _save(units, cur_unit, cur_dir, cur_prods, dates)
                cur_prods = []
                cur_dir = None
            if cur_unit and cur_unit in units:
                key = {'Потреблено': 'consumed', 'Вырабатано': 'produced',
                       'Дебаланс': 'imbalance', 'Дебаланс отн.': 'imbalance_rel'}.get(g)
                if key:
                    units[cur_unit]["summary"][key] = _read_days(ws, row_idx, num_days)
            continue

        if h and 'Суммарно' in str(h):
            if cur_unit and cur_dir and cur_prods:
                _save(units, cur_unit, cur_dir, cur_prods, dates)
                cur_prods = []
            continue

        if h and cur_unit and cur_dir:
            cur_prods.append({
                "product": str(h).strip(),
                "plan_month": _f(ws.cell(row=row_idx, column=COL_PLAN_MONTH).value),
                "plan_day": _f(ws.cell(row=row_idx, column=COL_PLAN_DAY).value),
                "days": _read_days(ws, row_idx, num_days),
            })

    if cur_unit and cur_dir and cur_prods:
        _save(units, cur_unit, cur_dir, cur_prods, dates)
    return units


def _read_days(ws, row_idx, num_days):
    m, r = [], []
    for i in range(num_days):
        m.append(_f(ws.cell(row=row_idx, column=DATA_START_COL + i * 2).value))
        r.append(_f(ws.cell(row=row_idx, column=DATA_START_COL + i * 2 + 1).value))
    return {"measured": m, "reconciled": r}


def _save(units, unit, direction, products, dates):
    rows = []
    for p in products:
        row = {"product": p["product"], "plan_month": p["plan_month"], "plan_day": p["plan_day"]}
        for i, d in enumerate(dates):
            ds = d.strftime("%Y-%m-%d")
            row[f"{ds}_meas"] = p["days"]["measured"][i] if i < len(p["days"]["measured"]) else 0.0
            row[f"{ds}_recon"] = p["days"]["reconciled"][i] if i < len(p["days"]["reconciled"]) else 0.0
        rows.append(row)
    df = pd.DataFrame(rows)
    key = "inputs" if direction == "Входящие" else "outputs"
    if units[unit][key] is None:
        units[unit][key] = df
    else:
        units[unit][key] = pd.concat([units[unit][key], df], ignore_index=True)


def _f(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


if __name__ == "__main__":
    f = os.path.join(os.path.dirname(__file__), "../../data/sample.xlsm")
    r = parse_report(f)
    print(f"Файл: {r['filename']}, Даты: {r['dates']}, Установки: {list(r['units'].keys())}")
    for name, d in r['units'].items():
        ni = len(d['inputs']) if d['inputs'] is not None else 0
        no = len(d['outputs']) if d['outputs'] is not None else 0
        print(f"  {name}: {ni} вх, {no} вых, дебаланс={d['summary']['imbalance']['measured']}")
